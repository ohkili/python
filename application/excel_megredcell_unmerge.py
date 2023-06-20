from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from copy import copy

# 병합된 셀을 분리하여 기존 데이터로 채우고 정렬, 포맷도 그대로 적용한다.

wb= load_workbook(filename= 'test.xlsx')

for sheet_name in wb.sheetnames:
        sheet= wb[sheet_name]
        mcr_coord_list= [mcr.coord for mcr in sheet.merged_cells.ranges]

        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row= range_boundaries(mcr)
            top_left_cell_value= sheet.cell(row=min_row, column=min_col).value
            top_left_cell_alignment= sheet.cell(row=min_row, column=min_col).alignment
            top_left_cell_format= sheet.cell(row=min_row, column=min_col).number_format
            sheet.unmerge_cells(mcr)
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value= top_left_cell_value
                    cell.alignment= copy(top_left_cell_alignment)
                    cell.number_format= copy(top_left_cell_format)

wb.save('merged_tmp.xlsx')